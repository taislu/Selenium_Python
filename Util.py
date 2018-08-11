# -*- coding: utf-8 -*-
"""
Created on Fri Aug 10 20:39:56 2018

@author: Tai
"""

from openpyxl import load_workbook
from openpyxl import Workbook
from datetime import datetime


class ReadExcel(object):
    
    def __init__(self, file_name, sheet_name):
        
        self.wb = load_workbook(file_name)
        self.sh = self.wb[sheet_name]
        
    def get_data(self, cell):
       
        return self.sh[cell].value
    
    
    
class WriteExcel(object):
    
    def __init__(self, file_name, sheet_name):
        
        self.wb = Workbook()
        self.ws1 = self.wb.active
        self.ws1.title = sheet_name
        self.filename = file_name
        
    def append_data(self, list):
        
        self.ws1.append(list)
        
    def save(self):
        
        self.wb.save(filename=self.filename)
        
def logInfo(msg):
    
    timestamp = datetime.now().isoformat()
    print(timestamp + " : " + msg)
    

def main():
    
    logInfo("Hello there!")
    
    f = "C:\\2018\\edureka\\selenium\\INPUT\\gmail_input.xlsx"
    sh = ReadExcel(f, 'Sheet1')
    print( sh.get_data('A1') ) 
    
    f1 = "C:\\2018\\edureka\\selenium\\INPUT\\gmail_output.xlsx"
    sh = WriteExcel(f1, 'Sheet1')
    L = [1, 2, 3, 4, 5]
    L1= ['A', 'B', 'C', 'D', 'E']
    sh.append_data(L)
    sh.append_data(L1)
    sh.save()


if __name__ == '__main__':
    main()